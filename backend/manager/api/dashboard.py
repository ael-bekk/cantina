from distutils.command.build_scripts import first_line_re
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import authentication, permissions, renderers
from rest_framework import status
from rest_framework import permissions
from urllib.parse import parse_qs
from django.http import HttpResponse
from django.contrib.auth.models import User
import datetime
import csv
import io
from itertools import groupby
from django.conf import settings
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from django.db.models import Count



from manager.models import Breakfast, Lunch, Dinner, Registry, Identity, Person
from .serializers import BreakfastSerializer, IdentitySerializer, LunchSerializer, DinnerSerializer, RegistrySerializer, UserSerializer, PersonSerializer
from ..utils import BREAKFAST, DINNER, LUNCH

from rest_framework.pagination import PageNumberPagination

class PersonPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 1000


class IsSuperUser(permissions.BasePermission):
    """
    Global permission check for superuser.
    """

    def has_permission(self, request, view):
        return request.user.is_superuser

    def has_object_permission(self, request, view, obj):
        return request.user.is_superuser


class DashboardMealsView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAdminUser]
    renderer_classes = [renderers.JSONRenderer]

    def solo_to_json(self, object):
        return {
            'name': object.name,
            'start_time': object.start_time,
            'end_time': object.end_time,
            'is_active': object.is_active,
        }

    def get(self, request):
        """
        Return the meals data
        """

        data = {
            'breakfast': self.solo_to_json(Breakfast.get_solo()),
            'lunch': self.solo_to_json(Lunch.get_solo()),
            'dinner': self.solo_to_json(Dinner.get_solo())
        }
        return Response(data, status=status.HTTP_200_OK)

    def patch(self, request, meal_name):
        if not meal_name in [BREAKFAST, LUNCH, DINNER]:
            return Response({'detail': 'Bad meal name.'}, status=status.HTTP_400_BAD_REQUEST)

        if meal_name == BREAKFAST:
            meal = Breakfast.get_solo()
            serializer = BreakfastSerializer(instance=meal, data=request.data)
        elif meal_name == LUNCH:
            meal = Lunch.get_solo()
            serializer = LunchSerializer(instance=meal, data=request.data)
        else:
            meal = Dinner.get_solo()
            serializer = DinnerSerializer(instance=meal, data=request.data)

        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.error, status=status.HTTP_400_BAD_REQUEST)


# generate xlsx file invoice
class DashboardMealsAnalyticsDownloadView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAdminUser]
    renderer_classes = [renderers.JSONRenderer]

    def get(self, request):
        queries = parse_qs(request.META['QUERY_STRING'])
        print(queries)
        if queries:
            return self.date_range_analytics(queries)

    def date_range_analytics(self, queries):
        if (not '__start_time' in list(queries.keys())) or (not '__end_time' in list(queries.keys())):
            return Response({'detail': "Invalid URL Queries."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            if queries['__start_time'][0] == 'null' or queries['__end_time'][0] == 'null':
                end_time = datetime.date.today() + datetime.timedelta(days=1)
                start_time = datetime.date.today() - datetime.timedelta(days=7)
            else:
                start_time = datetime.datetime.fromisoformat(
                    queries['__start_time'][0] + " 00:00:00")
                end_time = datetime.datetime.fromisoformat(
                    queries['__end_time'][0] + " 00:00:00")
        except Exception as e:
            print(e)
            return Response({'detail': "Invalid __start_time or __end_time format."}, status=status.HTTP_400_BAD_REQUEST)
        entries = RegistrySerializer(Registry.objects.filter(
            time__range=[start_time, end_time]), many=True).data

        # if len(entries) > 0:
        #     data = []
        #     for entry in entries:
        #         entry['time'] = entry['time'].split('T')[0]
        #     for key, value in groupby(entries, key=lambda entry: entry['time']):
        #         value = list(value)
        #         data.append({
        #             'date': key,
        #             'data': {
        #                     BREAKFAST: len([v for v in value if v['meal'] == BREAKFAST]),
        #                     LUNCH: len([v for v in value if v['meal'] == LUNCH]),
        #                     DINNER: len(
        #                         [v for v in value if v['meal'] == DINNER])
        #                     }
        #         })
        #     entries = data

        if len(entries) > 0:
            for entry in entries:
                entry["time"] = entry["time"].split("T")[0]

            # Create a dictionary to store counts for each date and meal type
            counts = {}
            for entry in entries:
                date = entry["time"]
                meal = entry["meal"]

                # Initialize counts for the date if not present
                if date not in counts:
                    counts[date] = {BREAKFAST: 0, LUNCH: 0, DINNER: 0}

                # Increment the count for the specific meal type
                counts[date][meal] += 1

            # Replace the original 'entries' with the updated 'counts'
            entries = [{"date": date, "data": counts[date]} for date in counts]


        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="analytics.xlsx"'
        workbook = Workbook()
        worksheet = workbook.active
        # cell width
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 20
        worksheet.row_dimensions[1].height = 40
        worksheet.row_dimensions[2].height = 40
        worksheet.row_dimensions[3].height = 40
        worksheet.row_dimensions[4].height = 30
        worksheet.append(['Articles Vendus\n\n Période du {} au {}\n\n Edité le {}'.format(start_time.strftime('%d/%m/%Y'), (end_time - datetime.timedelta(days=1)).strftime('%d/%m/%Y'), datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'))])
        # add date range
        # worksheet.append(['Période du {} au {}'.format(start_time.strftime('%d/%m/%Y'), end_time.strftime('%d/%m/%Y'))])
        worksheet.append([''])
        worksheet.append([''])
        worksheet.merge_cells('A1:D3')
        worksheet.title = 'Analytics'
        worksheet.append(['Date', 'Petit-déjeuner', 'Déjeuner', 'dîner'])
        # set bold
        for cell in worksheet[4]:
            cell.font = Font(bold=True)
        for index, entry in enumerate(entries):
            worksheet.append([entry['date'], entry['data'][BREAKFAST],
                            entry['data'][LUNCH], entry['data'][DINNER]])
            worksheet.row_dimensions[index + 5].height = 25
        # calculate total
        worksheet.append(['Total', '=SUM(B4:B{})'.format(len(entries) + 4), '=SUM(C4:C{})'.format(
            len(entries) + 4), '=SUM(D4:D{})'.format(len(entries) + 4)])
        # set total to 25 height
        worksheet.row_dimensions[len(entries) + 5].height = 30
        # set total bold
        for cell in worksheet[len(entries) + 5]:
            cell.font = Font(bold=True)
        # set border
        for row in worksheet.iter_rows(min_row=4, max_row=len(entries) + 5):
            for cell in row:
                cell.border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
        # first row bold
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        # align middle text
        for row in worksheet.iter_rows(min_row=1, max_row=len(entries) + 5):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        # color last cell
        worksheet['A{}'.format(len(entries) + 5)].fill = PatternFill(
            start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        worksheet['B{}'.format(len(entries) + 5)].fill = PatternFill(
            start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        worksheet['C{}'.format(len(entries) + 5)].fill = PatternFill(
            start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        worksheet['D{}'.format(len(entries) + 5)].fill = PatternFill(
            start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        workbook.save(response)

        return response

        # return Response({'entries': entries}, status=status.HTTP_200_OK)


class DashboardMealsAnalyticsView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAdminUser]
    renderer_classes = [renderers.JSONRenderer]

    def get(self, request):
        queries = parse_qs(request.META['QUERY_STRING'])
        if queries:
            return self.date_range_analytics(queries)
        # get today entries
        today_date = datetime.date.today()
        today_entries = Registry.objects.filter(
            time__day=today_date.day, time__month=today_date.month, time__year=today_date.year)
        today = {
            'breakfast': today_entries.filter(meal=BREAKFAST).count(),
            'lunch': today_entries.filter(meal=LUNCH).count(),
            'dinner': today_entries.filter(meal=DINNER).count()
        }
        # get yesterday entries
        yesterday_date = datetime.date.today() - datetime.timedelta(days=1)
        yesterday_entries = Registry.objects.filter(
            time__day=yesterday_date.day, time__month=yesterday_date.month, time__year=yesterday_date.year)
        yesterday = {
            'breakfast': yesterday_entries.filter(meal=BREAKFAST).count(),
            'lunch': yesterday_entries.filter(meal=LUNCH).count(),
            'dinner': yesterday_entries.filter(meal=DINNER).count()
        }
        return Response({'today': today, 'yesterday': yesterday}, status=status.HTTP_200_OK)

    def date_range_analytics(self, queries):
        if (not '__start_time' in list(queries.keys())) or (not '__end_time' in list(queries.keys())):
            return Response({'detail': "Invalid URL Queries."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            if queries['__start_time'][0] == 'null' or queries['__end_time'][0] == 'null':
                end_time = datetime.date.today() + datetime.timedelta(days=1)
                start_time = datetime.date.today() - datetime.timedelta(days=7)
            else:
                start_time = datetime.datetime.fromisoformat(
                    queries['__start_time'][0] + " 00:00:00")
                end_time = datetime.datetime.fromisoformat(
                    queries['__end_time'][0] + " 00:00:00")
        except Exception as e:
            print(e)
            return Response({'detail': "Invalid __start_time or __end_time format."}, status=status.HTTP_400_BAD_REQUEST)
        entries = RegistrySerializer(Registry.objects.filter(
            time__range=[start_time, end_time]), many=True).data
        # if len(entries) > 0:
        #     data = []
        #     for entry in entries:
        #         entry['time'] = entry['time'].split('T')[0]
        #     for key, value in groupby(entries, key=lambda entry: entry['time']):
        #         value = list(value)
        #         data.append({
        #             'date': key,
        #             'data': {
        #                     BREAKFAST: len([v for v in value if v['meal'] == BREAKFAST]),
        #                     LUNCH: len([v for v in value if v['meal'] == LUNCH]),
        #                     DINNER: len(
        #                         [v for v in value if v['meal'] == DINNER])
        #                     }
        #         })
        #     entries = data

        # code replaced by the code below
        if len(entries) > 0:
            for entry in entries:
                entry["time"] = entry["time"].split("T")[0]

            # Create a dictionary to store counts for each date and meal type
            counts = {}
            for entry in entries:
                date = entry["time"]
                meal = entry["meal"]

                # Initialize counts for the date if not present
                if date not in counts:
                    counts[date] = {BREAKFAST: 0, LUNCH: 0, DINNER: 0}

                # Increment the count for the specific meal type
                counts[date][meal] += 1
        # end code replaced by the code below
            return Response({'entries': [{"date": date, "data": counts[date]} for date in counts]}, status=status.HTTP_200_OK)
        return Response({'entries': []}, status=status.HTTP_200_OK)


class DashboardDownloadMealsAnalyticsView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAdminUser]
    renderer_classes = [renderers.JSONRenderer]

    def get(self, request):
        start_time = request.GET.get('__start_time', None)
        end_time = request.GET.get('__end_time', None)

        if not start_time or not end_time:
            return Response({'detail': "Missing required URL queries __start_time or __end_time."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            start_time = datetime.datetime.fromisoformat(
                start_time + " 00:00:00")
            end_time = datetime.datetime.fromisoformat(end_time + " 00:00:00")
        except Exception as e:
            print(e)
            return Response({'detail': "Invalid __start_time or __end_time format."}, status=status.HTTP_400_BAD_REQUEST)
        # entries = RegistrySerializer(Registry.objects.filter(time__range=[start_time, end_time]).order_by('-time'), many=True).data
        entries = Registry.objects.filter(
            time__range=[start_time, end_time]).order_by('-time')

        response = HttpResponse(
            content_type='text/csv',
            headers={'Content-Disposition': 'attachment; filename="comments.csv"'},
        )
        writer = csv.writer(response)
        writer.writerow(['Date', 'Meal', 'Cashier',
                         'Full Name', 'Login'])
        for entry in entries:
            writer.writerow([
                str(entry.time).split('.')[0],
                entry.meal,
                entry.cashier,
                entry.get_full_name(),
                entry.login
            ])

        return response


class DashboardManageUsersView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperUser]
    rendrrer_classes = [renderers.JSONRenderer]

    def get(self, request):
        users = User.objects.all().order_by('-id')
        serializer = UserSerializer(users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = UserSerializer(data=request.data)

        password1 = request.data.get("password1", None)
        password2 = request.data.get("password2", None)
        password_message = None
        if not password1 == password2:
            password_message = ['Passwords do not match.']
        if not password1 or not password2:
            password_message = ['Passwords cannot be empty.']

        if serializer.is_valid():
            if not password_message:
                if request.data['user_type'] == "admin":
                    user = User.objects.create_superuser(
                        username=request.data['username'], first_name=request.data['first_name'], last_name=request.data['last_name'], password=request.data['password1'])
                    return Response(UserSerializer(user).data, status=status.HTTP_200_OK)
                if request.data['user_type'] == "staff":
                    user = User.objects.create(
                        username=request.data['username'], first_name=request.data['first_name'], last_name=request.data['last_name'])
                    user.is_staff = True
                    user.set_password(password1)
                    user.save()
                    return Response(UserSerializer(user).data, status=status.HTTP_200_OK)
                else:
                    serializer.save()
                    user = User.objects.get(
                        username=serializer.data['username'])
                    user.set_password(request.data['password1'])
                    user.save()
                    return Response(serializer.data, status=status.HTTP_200_OK)
        errors = serializer.errors
        if password_message:
            errors.update({'password': password_message})
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request):
        try:
            if not request.data.get('id', None):
                raise ValueError("Missing required field 'id'")
            user = User.objects.get(id=request.data['id'])
        except Exception as e:
            print(e)
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        password1 = request.data.get("password1", None)
        password2 = request.data.get("password2", None)
        password_message = None
        if not password1 == password2:
            password_message = ['Passwords do not match.']

        serializer = UserSerializer(instance=user, data=request.data)
        if serializer.is_valid():
            if not password_message:
                serializer.save()
                if password1:
                    user.set_password(password1)
                    user.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
        errors = serializer.errors
        if password_message:
            errors.update({'password': password_message})
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request):
        try:
            if not request.data.get('id', None):
                raise ValueError("Missing required field 'id'")
            User.objects.get(id=request.data['id']).delete()
            return Response({}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class DashboardIdentityView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperUser]
    rendrrer_classes = [renderers.JSONRenderer]

    def get(self, request):
        identity = Identity.get_solo()
        return Response({'identity': identity.identity}, status=status.HTTP_200_OK)

    def patch(self, request):
        identity = Identity.get_solo()
        serializer = IdentitySerializer(instance=identity, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DashboardPersonView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperUser]
    # rendrrer_classes = [renderers.JSONRenderer]
    def get(self, request):
        kind = request.query_params.get('kind', None)
        if not kind:
            persons = Person.objects.all().order_by('authorized', '-id')
        else:
            persons = Person.objects.filter(kind=kind).order_by('authorized', '-id')
        paginator = PersonPagination()
        result_page = paginator.paginate_queryset(persons, request)
        serializer = PersonSerializer(result_page, many=True)
        return Response({
            'count': paginator.page.paginator.count,
            'num_pages': paginator.page.paginator.num_pages,
            'results': serializer.data,
            'staff': Person.objects.filter(kind='staff').count(),
            'student': Person.objects.filter(kind='student').count(),
            'prestataire': Person.objects.filter(kind='prestataire').count(),
            'pooler': Person.objects.filter(kind='pooler').count(),
        })


    def post(self, request):

        def check_or_create_persons(data, create=False):
            errors = []
            for row in data:
                if all(x in list(row.keys()) for x in ['full_name', 'login', 'badge', 'kind', 'authorized', 'superuser']):
                    person = {
                        "full_name": str(row['full_name']).capitalize(),
                        "login": str(row['login']).lower(),
                        "badge": str(row['badge']).lstrip("0"),
                        "kind": str(row['kind']).lower(),
                        "authorized": bool(str(row['authorized']).lower()),
                        "superuser": True if str(row['superuser']).lower() == 'true' else False,
                    }
                    serializer = PersonSerializer(data=person)
                    if serializer.is_valid():
                        if create:
                            serializer.save()
                    else:
                        print(serializer.errors)
                        for key, value in serializer.errors.items():
                            errors.append(f'LINE {data.line_num}: COLUMN "{key}" ERROR "{list(value)[0]}"')
                else:
                    errors.append(f'LINE {data.line_num}: Missing required column')
            return errors
        file = request.data.get('file', None)
        if file:
            file = file.read().decode('utf-8')
            read_for_check = csv.DictReader(io.StringIO(file))
            read_for_create = csv.DictReader(io.StringIO(file))
            errors = check_or_create_persons(read_for_check)
            if errors:
                return Response({"errors": errors}, status=status.HTTP_400_BAD_REQUEST)
            else:
                check_or_create_persons(read_for_create, create=True)
                return Response({}, status=status.HTTP_200_OK)
        else:
            serializer = PersonSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request):
        try:
            if not request.data.get('id', None):
                raise ValueError("Missing required field 'id'")
            person = Person.objects.get(id=request.data['id'])
        except Exception as e:
            print(e)
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        serializer = PersonSerializer(instance=person, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request):
        ids = request.data
        for id in ids:
            try:
                Person.objects.get(id=id).delete()
            except Person.DoesNotExist as e:
                return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({}, status=status.HTTP_200_OK)

class DashboardPersonViewSearch(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperUser]
    # rendrrer_classes = [renderers.JSONRenderer]

    # search inside persons by full_name, login, badge
            # personies = Person.objects.filter(Q(full_name__icontains=search) | Q(login__icontains=search) | Q(badge__icontains=search)).order_by('authorized', '-id')
    def get(self, request):
        search = request.query_params.get('search', None)
        if not search:
            persons = Person.objects.all().order_by('authorized', '-id')
        else:
            persons = Person.objects.filter(Q(full_name__icontains=search) | Q(login__icontains=search) | Q(badge__icontains=search)).order_by('authorized', '-id')
        paginator = PersonPagination()
        result_page = paginator.paginate_queryset(persons, request)
        serializer = PersonSerializer(result_page, many=True)
        return Response({
            'count': paginator.page.paginator.count,
            'num_pages': paginator.page.paginator.num_pages,
            'results': serializer.data,
            'staff': Person.objects.filter(kind='staff').count(),
            'student': Person.objects.filter(kind='student').count(),
            'prestataire': Person.objects.filter(kind='prestataire').count(),
            'pooler': Person.objects.filter(kind='pooler').count(),
        })


class DashboardRecordsSearchView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperUser]

    def get(self, request):
        login = request.GET.get('login', None)
        start_date = request.GET.get('start_date', None)
        end_date = request.GET.get('end_date', None)

        if not login:
            return Response({'detail': "Missing required URL queries &login"}, status=status.HTTP_400_BAD_REQUEST)

        records = Registry.objects.filter(login=login).order_by('-time')

        if start_date and start_date != 'null' and end_date and end_date != 'null':
            try:
                start_date = (datetime.datetime.fromisoformat(
                    start_date + " 00:00:00")).strftime("%Y-%m-%d")
                end_date = (datetime.datetime.fromisoformat(
                    end_date + " 00:00:00") + datetime.timedelta(days=2)).strftime("%Y-%m-%d")
                records = records.filter(time__range=[start_date, end_date])
            except Exception as e:
                print(e)
                return Response({'detail': "Invalid start_date or end_date format."}, status=status.HTTP_400_BAD_REQUEST)

        data = []
        for record in records:
            data.append({
                'time': str(record.time).split('.')[0],
                'login': record.login,
                'full_name': record.get_full_name(),
                'meal': record.meal,
                'cashier': record.cashier,
            })
        return Response(data, status=status.HTTP_200_OK)


class DashboardRecordsSearchEagleEyeView(APIView):
    authentication_classes = ()
    permission_classes = ()

    def get(self, request):
        token = request.headers.get('X-Secret', None)
        
        if not token:
            return Response({'detail': "No authorization token provided"}, status=status.HTTP_401_UNAUTHORIZED)

        if token != settings.EAGLE_EYE_TOKEN:
            return Response({'detail': "Invalid authorization token"}, status=status.HTTP_401_UNAUTHORIZED)
        
        # login = request.GET.get('login', None)
        login = request.GET.get('login', None)
        start_date = request.GET.get('start_date', None)
        end_date = request.GET.get('end_date', None)

        if not start_date or not end_date:
            return Response({'detail': "Invalid start_date or end_date format."}, status=status.HTTP_400_BAD_REQUEST)

        if not login:
            try:
                start_date = (datetime.datetime.fromisoformat(
                    start_date + " 00:00:00") + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
                end_date = (datetime.datetime.fromisoformat(
                    end_date + " 00:00:00") + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
            except Exception as e:
                print(e)
                return Response({'detail': "Invalid start_date or end_date format."}, status=status.HTTP_400_BAD_REQUEST)
            # records = Registry.objects.order_by('-time')
            records = Registry.objects.all().values('login').filter(time__range=[start_date, end_date]).annotate(Count('login'))
            # aaammari


            # data = []
            # for record in records:
            #     data.append(
            #         record.login,
            #     )
            # count each login how many meals they ate
            # res = {}
            # for i in data:
            #     res[i] = data.count(i)
            return Response(records, status=status.HTTP_200_OK)
        else:
            records = Registry.objects.filter(login=login).order_by('-time')

            if start_date and start_date != 'null' and end_date and end_date != 'null':
                try:
                    start_date = (datetime.datetime.fromisoformat(
                        start_date + " 00:00:00") + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
                    end_date = (datetime.datetime.fromisoformat(
                        end_date + " 00:00:00") + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
                    records = records.filter(time__range=[start_date, end_date])
                    # records lenght
                    return Response({
                        'count': records.count(),
                    }, status=status.HTTP_200_OK)
                except Exception as e:
                    print(e)
                    return Response({'detail': "Invalid start_date or end_date format."}, status=status.HTTP_400_BAD_REQUEST)
            
            return Response(data, status=status.HTTP_200_OK)